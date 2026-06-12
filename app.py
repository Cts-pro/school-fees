import os
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.secret_key = 'super_secret_permanent_key_123'
EXCEL_FILE = '/data/school_fees_by_class.xlsx'

font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
font_body = Font(name="Calibri", size=11)
font_locked = Font(name="Calibri", size=11, italic=True, color="595959")
fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
fill_locked = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
fill_pending = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

def init_excel():
    dirname = os.path.dirname(EXCEL_FILE)
    if dirname and not os.path.exists(dirname):
        os.makedirs(dirname, exist_ok=True)
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Grade 5"
        ws.views.sheetView[0].showGridLines = True
        headers = ["Record ID", "Student Name", "Category", "Particulars List", "Quantities List", "Amounts List", "Total Amount", "Status", "Is Locked", "Invoice Number"]
        ws.append(headers)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center")
        wb.save(EXCEL_FILE)

def get_all_records_from_excel():
    if not os.path.exists(EXCEL_FILE):
        init_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    all_records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value is None:
                continue
            try:
                all_records.append({
                    "id": int(ws.cell(row=row, column=1).value),
                    "name": ws.cell(row=row, column=2).value,
                    "category": str(ws.cell(row=row, column=3).value or "Fees").strip(),
                    "particulars": str(ws.cell(row=row, column=4).value or ""),
                    "quantities": str(ws.cell(row=row, column=5).value or ""),
                    "amounts": str(ws.cell(row=row, column=6).value or ""),
                    "total_amount": float(ws.cell(row=row, column=7).value or 0.0),
                    "class": sheet_name,
                    "status": ws.cell(row=row, column=8).value,
                    "is_locked": int(ws.cell(row=row, column=9).value or 0),
                    "invoice_num": str(ws.cell(row=row, column=10).value or "DRAFT-UNASSIGNED")
                })
            except Exception:
                continue
    return all_records

def get_next_invoice_number_for_category(category):
    records = get_all_records_from_excel()
    count = 0
    for r in records:
        if r['category'] == category and r['is_locked'] == 1:
            count += 1
    prefix = {"Fees": "Tuition-", "Books": "Book-", "Uniform": "Uniform-"}
    return f"{prefix.get(category, 'INV-')}{count + 1:04d}"

@app.route('/')
def index():
    if 'user_role' not in session:
        return redirect(url_for('login'))
    records = get_all_records_from_excel()
    return render_template('index.html', records=records, role=session.get('user_role'), username=session.get('username'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if username == 'admin' and password == 'admin123':
            session['user_role'] = 'admin'
            session['username'] = 'Admin'
            return redirect(url_for('index'))
        elif username == 'staff' and password == 'staff123':
            session['user_role'] = 'user'
            session['username'] = 'Bursar Staff'
            return redirect(url_for('index'))
        else:
            flash('Invalid login credentials.')
    return render_template('index.html', login_page=True)

@app.route('/add_student', methods=['POST'])
def add_student():
    if 'user_role' not in session:
        return redirect(url_for('login'))
        
    student_name = request.form['student_name']
    student_class = request.form['student_class'].strip()
    category = request.form['category'].strip()
    particulars = request.form.get('particulars', '')
    quantities = request.form.get('quantities', '')
    amounts_str = request.form.get('amounts', '')
    
    if not particulars.strip():
        particulars = "Default Item"
    if not quantities.strip():
        quantities = "1"
    if not amounts_str.strip():
        amounts_str = "0.00"
    
    try:
        amounts_list = [float(x.strip() if x.strip() else 0) for x in amounts_str.split(',')]
        total_amount = sum(amounts_list)
    except ValueError:
        total_amount = 0.0
        
    wb = openpyxl.load_workbook(EXCEL_FILE)
    if student_class not in wb.sheetnames:
        ws = wb.create_sheet(title=student_class)
        ws.views.sheetView[0].showGridLines = True
        headers = ["Record ID", "Student Name", "Category", "Particulars List", "Quantities List", "Amounts List", "Total Amount", "Status", "Is Locked", "Invoice Number"]
        ws.append(headers)
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col).font = font_header
            ws.cell(row=1, column=col).fill = fill_header
    else:
        ws = wb[student_class]
        
    all_records = get_all_records_from_excel()
    next_id = max([r['id'] for r in all_records]) + 1 if all_records else 1
    
    ws.append([next_id, student_name, category, particulars, quantities, amounts_str, total_amount, "Pending", 0, "DRAFT-UNASSIGNED"])
    
    new_row_idx = ws.max_row
    for col in range(1, 11):
        cell = ws.cell(row=new_row_idx, column=col)
        cell.fill = fill_pending
        cell.font = font_body
        if col == 7:
            cell.number_format = 'Rs#,##0.00'
            
    wb.save(EXCEL_FILE)
    flash(f"Draft matrix saved successfully for {student_name}!")
    return redirect(url_for('index'))

@app.route('/edit_student', methods=['POST'])
def edit_student():
    if 'user_role' not in session:
        return redirect(url_for('login'))
        
    row_id = int(request.form['record_id'])
    sheet_class = request.form['sheet_class']
    updated_name = request.form['edit_name']
    updated_class = request.form['edit_class'].strip()
    
    particulars_list = request.form.getlist('edit_particulars')
    quantities_list = request.form.getlist('edit_quantities')
    amounts_list = request.form.getlist('edit_amounts')
    
    parts_str = ",".join(particulars_list)
    qtys_str = ",".join(quantities_list)
    ams_str = ",".join(amounts_list)
    
    try:
        total_amount = sum([float(x if str(x).strip() else 0) for x in amounts_list])
    except ValueError:
        flash("Numeric formatting handling failed.")
        return redirect(url_for('index'))
        
    wb = openpyxl.load_workbook(EXCEL_FILE)
    if sheet_class in wb.sheetnames:
        ws = wb[sheet_class]
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == row_id:
                if ws.cell(row=row, column=9).value == 1 and session.get('user_role') != 'admin':
                    flash("Action Blocked: Row is finalized.")
                    return redirect(url_for('index'))
                
                ws.cell(row=row, column=2).value = updated_name
                ws.cell(row=row, column=4).value = parts_str
                ws.cell(row=row, column=5).value = qtys_str
                ws.cell(row=row, column=6).value = ams_str
                ws.cell(row=row, column=7).value = total_amount
                
                current_lock = ws.cell(row=row, column=9).value
                inv_num = ws.cell(row=row, column=10).value
                category = ws.cell(row=row, column=3).value
                status = ws.cell(row=row, column=8).value
                
                if updated_class != sheet_class:
                    ws.delete_rows(row, 1)
                    if updated_class not in wb.sheetnames:
                        new_ws = wb.create_sheet(title=updated_class)
                        new_ws.views.sheetView[0].showGridLines = True
                        headers = ["Record ID", "Student Name", "Category", "Particulars List", "Quantities List", "Amounts List", "Total Amount", "Status", "Is Locked", "Invoice Number"]
                        new_ws.append(headers)
                    else:
                        new_ws = wb[updated_class]
                        
                    new_ws.append([row_id, updated_name, category, parts_str, qtys_str, ams_str, total_amount, status, current_lock, inv_num])
                    target_row = new_ws.max_row
                    for c in range(1, 11):
                        cell = new_ws.cell(row=target_row, column=c)
                        if current_lock == 1:
                            cell.fill = fill_locked
                            cell.font = font_locked
                        else:
                            cell.fill = fill_pending
                            cell.font = font_body
                else:
                    for c in range(1, 11):
                        cell = ws.cell(row=row, column=c)
                        if current_lock == 1:
                            cell.fill = fill_locked
                            cell.font = font_locked
                        else:
                            cell.fill = fill_pending
                            cell.font = font_body
                break
                
    wb.save(EXCEL_FILE)
    flash("Changes committed to master matrix successfully.")
    return redirect(url_for('index'))

@app.route('/delete_student/<sheet_name>/<int:row_id>', methods=['POST'])
def delete_student(sheet_name, row_id):
    if session.get('user_role') != 'admin':
        return redirect(url_for('login'))
        
    wb = openpyxl.load_workbook(EXCEL_FILE)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == row_id:
                ws.delete_rows(row, 1)
                wb.save(EXCEL_FILE)
                flash("Permanently removed record row.")
                break
    return redirect(url_for('index'))

@app.route('/generate_invoice/<sheet_name>/<int:row_id>')
def generate_invoice(sheet_name, row_id):
    if 'user_role' not in session:
        return redirect(url_for('login'))
    wb = openpyxl.load_workbook(EXCEL_FILE)
    record_data = {}
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == row_id:
                category = str(ws.cell(row=row, column=3).value or "Fees").strip()
                is_locked = int(ws.cell(row=row, column=9).value or 0)
                
                if is_locked == 0:
                    assigned_invoice = get_next_invoice_number_for_category(category)
                    ws.cell(row=row, column=8).value = "Invoiced"
                    ws.cell(row=row, column=9).value = 1
                    ws.cell(row=row, column=10).value = assigned_invoice
                    for col in range(1, 11):
                        ws.cell(row=row, column=col).fill = fill_locked
                        ws.cell(row=row, column=col).font = font_locked
                else:
                    assigned_invoice = str(ws.cell(row=row, column=10).value)
                    
                record_data = {
                    "id": row_id,
                    "name": ws.cell(row=row, column=2).value,
                    "category": category,
                    "particulars": str(ws.cell(row=row, column=4 Rely on default if empty" or "Items")).split(','),
                    "quantities": str(ws.cell(row=row, column=5).value or "1").split(','),
                    "amounts": str(ws.cell(row=row, column=6).value or "0").split(','),
                    "total": float(ws.cell(row=row, column=7).value or 0.0),
                    "class": sheet_name,
                    "invoice_num": assigned_invoice
                }
                wb.save(EXCEL_FILE)
                break

    table_rows_html = ""
    for i in range(len(record_data['particulars'])):
        part = record_data['particulars'][i].strip()
        qty = record_data['quantities'][i].strip() if i < len(record_data['quantities']) else "1"
        line_total = float(record_data['amounts'][i].strip() if record_data['amounts'][i].strip() else 0)
        rate = line_total / float(qty) if float(qty) > 0 else 0.0
        table_rows_html += f"<tr><td class='text-center'>{i+1}</td><td>{part}</td><td class='text-center'>{qty}</td><td class='text-end'>Rs {rate:,.2f}</td><td class='text-end'>Rs {line_total:,.2f}</td></tr>"

    return f"""
    <html>
    <head>
        <title>Invoice - {record_data['name']}</title>
        <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css">
        <style>
            body {{ background: #fdfdfd; font-family: Arial, sans-serif; color: #000; }}
            .outer-container {{ max-width: 850px; margin: 20px auto; padding: 20px; border: 1px solid #ccc; background: #fff; }}
            .header-border {{ border: 1px solid #000; padding: 20px; }}
            .school-title {{ font-size: 24px; font-weight: bold; text-transform: uppercase; }}
            .trust-title {{ font-size: 12px; font-weight: bold; margin-bottom: 5px; color: #333; }}
            .table-invoice th, .table-invoice td {{ border: 1px solid #000 !important; font-size: 14px; padding: 6px; }}
            .table-invoice th {{ background-color: #f2f2f2 !important; }}
            .d-print-none button, .d-print-none a {{ width: 100% !important; }}
        </style>
    </head>
    <body>
        <div class="outer-container">
            <div class="header-border">
                <div class="row align-items-center mb-3">
                    <div class="col-3 text-center">
                        <div style="border: 1px solid #333; padding: 15px; font-size: 11px; font-weight: bold;">LOGO PLACEHOLDER</div>
                    </div>
                    <div class="col-9 text-center" style="padding-right: 50px;">
                        <div class="school-title">VINAYAKA HIGH SCHOOL</div>
                        <div class="trust-title">SRI ANNAPOORNESHWARI EDUCATION TRUST R</div>
                        <div class="meta-info">Taripura - 571415, Srirangapatna Taluk, Mandya District</div>
                        <div class="meta-info"><strong>Mob:</strong> 8971577685</div>
                    </div>
                </div>
                <hr style="border-top: 1px solid #000; margin: 15px 0;">
                <div class="row mb-4" style="font-size: 15px; line-height: 1.6;">
                    <div class="col-6">
                        <div><strong>Invoice #:</strong> {record_data['invoice_num']}</div>
                        <div><strong>Invoice date:</strong> 11-06-2026</div>
                        <div><strong>Account Category:</strong> {record_data['category']}</div>
                    </div>
                    <div class="col-6" style="padding-left: 80px;">
                        <div><strong>Name:</strong> {record_data['name']}</div>
                        <div><strong>Class:</strong> {record_data['class']}</div>
                        <div><strong>Amount Due:</strong> Rs {record_data['total']:,.2f}</div>
                    </div>
                </div>
                <table class="table table-bordered table-invoice mb-0">
                    <thead><tr><th class="text-center" style="width: 70px;">SL. No</th><th>Particulars Description</th><th class="text-center" style="width: 80px;">Qty</th><th class="text-end" style="width: 130px;">Rate</th><th class="text-end" style="width: 140px;">Amount</th></tr></thead>
                    <tbody>
                        {table_rows_html}
                        <tr class="fw-bold"><td colspan="4" class="text-end">Total</td><td class="text-end">Rs {record_data['total']:,.2f}</td></tr>
                    </tbody>
                </table>
                <div class="row align-items-end mt-5" style="min-height: 80px;"><div class="col-12 text-end" style="font-size: 14px; padding-right: 30px;"><div>Signature</div></div></div>
            </div>
            <div class="d-print-none mt-4 d-flex gap-2"><button class="btn btn-primary w-100 fw-bold" onclick="window.print()">Print This Invoice</button><a href="/" class="btn btn-outline-secondary w-100">Return to Portal</a></div>
        </div>
    </body>
    </html>
    """

@app.route('/download_database')
def download_database():
    if 'user_role' not in session:
        return redirect(url_for('login'))
    if os.path.exists(EXCEL_FILE):
        return send_file(EXCEL_FILE, as_attachment=True, download_name="vinayaka_school_database.xlsx")
    return redirect(url_for('index'))

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

init_excel()

if __name__ == '__main__':
    app.run(debug=False)
